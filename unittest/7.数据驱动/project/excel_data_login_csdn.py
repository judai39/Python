import unittest
from time import sleep
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from ddt import ddt,data,unpack
import openpyxl as openpyxl
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# 读取excel中的信息,使用xlrd,openpyxl
def read_excel():
    xlsx=openpyxl.load_workbook("unittest/7.数据驱动/project/account_message.xlsx")
    sheet1=xlsx['Sheet1']
    print(sheet1.max_row)
    print(sheet1.max_column)
    print('=======================================================')
    allList=[]
    for row in range(2,sheet1.max_row+1):
        # 每次都会新建一个rowlist,对的,最后需要把所有
        rowlist=[]
        for column in range(1,sheet1.max_column+1):
            rowlist.append(sheet1.cell(row,column).value)
        allList.append(rowlist)
    return allList

# 
@ddt
class ExcelText(unittest.TestCase):
    def setUp(self):
        self.driver=webdriver.Chrome(service=Service('C:/Users/judai/.chromedriver/chromedriver143/chromedriver.exe'))
        self.driver.get("https://www.bilibili.com")
        self.driver.maximize_window()
    @data(read_excel()[0])
    def test_bilibili_login(self,account_message_list):
        login_span=WebDriverWait(self.driver,5).until(
            EC.presence_of_element_located((By.XPATH,"//div[@class='bili-header__bar']/ul[@class='right-entry']/li[1]//div[@class='header-login-entry']/span[1]"))
        )
        action_chains=ActionChains(self.driver)
        action_chains.move_to_element(login_span).perform()
        login_button=WebDriverWait(self.driver,10).until(
            EC.presence_of_element_located((By.XPATH,"//div[@class='v-popover-content']//div[@class='login-btn']"))
        )
        action_chains.move_to_element(login_button).perform()
        action_chains.click(login_button).perform()
        WebDriverWait(self.driver,5).until(
            EC.presence_of_element_located((By.XPATH,"//form[@class='tab__form']"))
        )
        account_input=self.driver.find_element(By.XPATH,"//form[@class='tab__form']/div[1]/input[1]")
        password_input=self.driver.find_element(By.XPATH,"//form[@class='tab__form']/div[3]/input[1]")
        account_input.send_keys(account_message_list[0])
        password_input.send_keys(account_message_list[1])
    
    def tearDown(self) -> None:
        sleep(3)
        self.driver.quit()
    
        
if __name__=="__main__":
    unittest.main()